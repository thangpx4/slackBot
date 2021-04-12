import logging

import openpyxl

# Give the location of the file

try:
    path = "../static/demo.xlsx"

    # To open the workbook
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)

    # Get workbook active sheet object
    # from the active attribute
    sheet_obj = wb_obj.active

    maxRow = sheet_obj.max_row  # get how many rows in sheet
    sheet_obj = wb_obj.active

    max_col = sheet_obj.max_column  # get how many columns in sheet
    listGroup = list()
except Exception as e:
    logging.basicConfig(filename='../static/getExcel.log', filemode='a+', format='%(name)s - %(levelname)s - %(message)s')
    logging.warning(e)
    print("getExcel.py expection: ", e)
def getListTopic():
    # get all name's topic in excel file
    listTopic = list()
    for i in range(2, max_col + 1):
        cell_obj6 = sheet_obj.cell(row=1, column=i)
        if cell_obj6.value == "Topic":
            for j in range(2, maxRow + 1):
                cell_obj7 = sheet_obj.cell(row=j, column=int(cell_obj6.col_idx))
                if cell_obj7.value:
                    listTopic.append((cell_obj7.value).strip())
    return listTopic


def getListGroup():
    # get all name's group in excel file
    listGroup = list()
    for i in range(1, max_col + 1):
        cell_obj6 = sheet_obj.cell(row=1, column=i)
        if cell_obj6.value == "Group":
            for j in range(2, maxRow + 1):
                cell_obj7 = sheet_obj.cell(row=j, column=int(cell_obj6.col_idx))
                if cell_obj7.value:
                    listGroup.append(str(cell_obj7.value))
    return listGroup

if __name__ == '__main__':
    getListGroup()
    getListTopic()
